#!/usr/bin/env python3
"""
FPT Chat ASM Report Tool
Phân tích báo cáo hàng ngày của ASM từ lịch sử chat FPT Chat.

Usage:
    python fpt_chat_stats.py --today
    python fpt_chat_stats.py --today --excel bao_cao.xlsx
    python fpt_chat_stats.py --from 2026-04-01 --to 2026-04-16
    python fpt_chat_stats.py --save raw.json
    python fpt_chat_stats.py --load raw.json --today
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from datetime import date as _date, datetime, timedelta, time, timezone
from typing import Literal

ReportType = Literal["daily_shop_vt", "weekend_tttc"]

try:
    import requests
except ImportError:
    print("Thiếu thư viện 'requests'. Chạy: pip install requests", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(path: str) -> dict:
    """Đọc config JSON. Trả về {} nếu file không tồn tại."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: config file '{path}' không hợp lệ: {e}", file=sys.stderr)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_group_id(value: str) -> str:
    """Trích group ID từ URL hoặc trả về nguyên nếu đã là ID."""
    match = re.search(r'([a-f0-9]{24})', value)
    return match.group(1) if match else value


def parse_dt(iso: str) -> datetime | None:
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except Exception:
        return None


def parse_date_arg(value: str, end_of_day: bool = False) -> datetime:
    """Chuyển 'YYYY-MM-DD' thành datetime UTC có timezone."""
    try:
        d = datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        print(f"Error: định dạng ngày không hợp lệ '{value}'. Dùng YYYY-MM-DD.", file=sys.stderr)
        sys.exit(1)
    t = time(23, 59, 59) if end_of_day else time(0, 0, 0)
    return datetime.combine(d, t, tzinfo=timezone.utc)


def to_vn_str(iso: str) -> str:
    """Chuyển ISO datetime thành chuỗi giờ VN (UTC+7)."""
    dt = parse_dt(iso)
    if not dt:
        return iso
    vn = datetime.fromtimestamp(dt.timestamp() + 7 * 3600, tz=timezone.utc)
    return vn.strftime("%Y-%m-%d %H:%M:%S")


def fmt_date_range(date_from: datetime | None, date_to: datetime | None) -> str:
    if not date_from and not date_to:
        return "Toàn bộ lịch sử"
    from_str = date_from.strftime("%Y-%m-%d") if date_from else "..."
    to_str = date_to.strftime("%Y-%m-%d") if date_to else "..."
    return f"{from_str} → {to_str}"


def build_session(token: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    })
    session.cookies.set("fchat_ddtk", token, domain="api-chat.fpt.com")
    return session


def filter_by_date(messages: list,
                   date_from: datetime | None,
                   date_to: datetime | None) -> list:
    if not date_from and not date_to:
        return messages
    result = []
    for m in messages:
        dt = parse_dt(m.get("createdAt", ""))
        if dt is None:
            result.append(m)
            continue
        if date_from and dt < date_from:
            continue
        if date_to and dt > date_to:
            continue
        result.append(m)
    return result


# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def fetch_page(session: requests.Session, base_url: str, group_id: str,
               limit: int, before_inc=None) -> dict:
    url = f"{base_url}/message-query/group/{group_id}/message"
    params: dict = {"limit": limit}
    if before_inc is not None:
        params["messageIdInc"] = before_inc
        params["cursorType"] = "PREVIOUS"
    resp = session.get(url, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def fetch_group_info(session: requests.Session, base_url: str, group_id: str) -> dict:
    """Lấy thông tin nhóm (name, ...). Trả {} nếu lỗi hoặc endpoint không tồn tại."""
    try:
        resp = session.get(f"{base_url}/group-management/group/{group_id}", timeout=10)
        if resp.ok:
            return resp.json()
    except Exception:
        pass
    return {}


def fetch_group_members(session: requests.Session, base_url: str,
                        group_id: str, limit: int = 50) -> list:
    """Lấy toàn bộ thành viên group qua phân trang page-based."""
    members = []
    page = 1
    print("[*] Fetching group members ...", file=sys.stderr)
    while True:
        url = f"{base_url}/group-management/group/{group_id}/participant"
        try:
            resp = session.get(url, params={"limit": limit, "page": page}, timeout=30)
            resp.raise_for_status()
        except Exception as e:
            print(f"  [!] Lỗi fetch group members (page {page}): {e}", file=sys.stderr)
            return []
        data = resp.json()
        items = data if isinstance(data, list) else data.get(
            "data", data.get("members", data.get("participants", data.get("items", [])))
        )
        if not items:
            break
        members.extend(items)
        if len(items) < limit:
            break
        page += 1
    print(f"[✓] {len(members)} thành viên", file=sys.stderr)
    return members


def fetch_all_messages(token: str, group_id: str,
                       base_url: str = "https://api-chat.fpt.com",
                       limit: int = 50,
                       date_from: datetime | None = None) -> list:
    """Lấy toàn bộ tin nhắn bằng cách phân trang ngược (oldest-first)."""
    session = build_session(token)
    all_messages = []
    before_inc = None
    page = 0

    print(f"[*] Fetching messages từ group {group_id} ...", file=sys.stderr)

    while True:
        page += 1
        data = fetch_page(session, base_url, group_id, limit, before_inc)
        regulars = data.get("regulars", [])

        if not regulars:
            break

        all_messages = regulars + all_messages
        fetched = len(regulars)
        oldest_dt = parse_dt(regulars[0].get("createdAt", ""))
        oldest_str = oldest_dt.strftime("%Y-%m-%d") if oldest_dt else "?"
        print(f"  Page {page}: +{fetched} messages (total: {len(all_messages)}, oldest: {oldest_str})",
              file=sys.stderr)

        if fetched < limit:
            break

        if date_from and oldest_dt and oldest_dt < date_from:
            print(f"  [✓] Đã đủ dữ liệu từ {date_from.strftime('%Y-%m-%d')}, dừng fetch.",
                  file=sys.stderr)
            break

        before_inc = regulars[0]["messageIdInc"]

    print(f"[✓] Tổng: {len(all_messages)} messages", file=sys.stderr)
    return all_messages


# ---------------------------------------------------------------------------
# Weekly Report Classifier
# ---------------------------------------------------------------------------

WEEKLY_SCORE_THRESHOLD = 3
WEEKLY_MIN_LENGTH = 150

_WEEKLY_RE_UNIT = re.compile(
    r"\b(tttc|vx\s|shop\b|trung\s*tâm|chi\s*nhánh|lc\s+hcm)",
    re.IGNORECASE,
)
_WEEKLY_RE_METRIC = re.compile(
    r"\d+\s*%|\d+(?:\.\d+)?\s*(tr|triệu|m\b|k\b|đ\b|cọc|bill|khách|lượt|gói)",
    re.IGNORECASE,
)
_WEEKLY_RE_OPEN_CLOSE = re.compile(
    r"(đánh\s*giá|báo\s*cáo|em\s+(?:xin\s+)?cảm\s*ơn|dạ\s+em\s+(?:gửi|bc))",
    re.IGNORECASE,
)
_WEEKLY_RE_SECTION = re.compile(
    r"(?m)^\s*[-–•\d\.]*\s*(kết\s*quả|tích\s*cực|vấn\s*đề|đã\s*làm|ngày\s*mai"
    r"|giải\s*pháp|hành\s*động|tổng\s*quan|phân\s*tích)\s*[:：]",
    re.IGNORECASE,
)


def _score_weekly_message(content: str) -> int:
    """Tính điểm phân loại báo cáo tuần (0..6) cho một đoạn text."""
    if not content:
        return 0
    flags = (
        len(content.strip()) >= WEEKLY_MIN_LENGTH,
        "\n" in content,
        bool(_WEEKLY_RE_UNIT.search(content)),
        bool(_WEEKLY_RE_METRIC.search(content)),
        bool(_WEEKLY_RE_OPEN_CLOSE.search(content)),
        bool(_WEEKLY_RE_SECTION.search(content)),
    )
    return sum(1 for f in flags if f)


# ---------------------------------------------------------------------------
# ASM Report Analysis
# ---------------------------------------------------------------------------

_REPORT_KEYWORDS = (
    "shop", "tttc", "vx hcm", "coc",
    "doanh thu", "dt %", "hot", "ra tiem",
    "tvv", "tu van", "kh", "bill",
)


def _strip_diacritics(s: str) -> str:
    """Lower + strip Vietnamese diacritics for keyword matching tolerance.

    `đ`/`Đ` handled explicitly because NFKD doesn't decompose them — they're
    base letters (U+0111 / U+0110), not composed of d + combining mark.
    """
    lowered = s.lower().replace("đ", "d")
    nfkd = unicodedata.normalize("NFKD", lowered)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def detect_report_candidates(messages: list) -> list:
    """L2 heuristic pre-filter: length ≥ 80 + ≥ 2 digits + ≥ 1 keyword.

    Cheap signal — LLM extraction phía sau quyết định loại + parse fields.
    Diacritic-insensitive keyword match (user gõ thiếu dấu vẫn pass).
    """
    out = []
    digit_re = re.compile(r"\d")
    for msg in messages:
        if msg.get("type") != "TEXT":
            continue
        content = msg.get("content") or ""
        if len(content) < 80:
            continue
        if len(digit_re.findall(content)) < 2:
            continue
        normalized = _strip_diacritics(content)
        if not any(kw in normalized for kw in _REPORT_KEYWORDS):
            continue
        out.append(msg)
    return out


def extract_all_reports(messages: list) -> list:
    """Pre-filter messages, then LLM-extract each candidate.

    Runs LLM calls concurrently using a bounded ThreadPoolExecutor
    (size via env `LLM_MAX_WORKERS`, default 4 — sized for OpenAI tier 1
    ~200k TPM). Order is preserved: executor.map yields results in input
    order regardless of completion order.

    Returns a flat list of Report dicts (may be longer than candidates because
    one message can produce multiple reports; may contain unparseable stubs
    with parse_error set)."""
    import llm_extractor
    candidates = detect_report_candidates(messages)
    if not candidates:
        return []
    workers = min(llm_extractor._read_max_workers(), len(candidates))
    out: list = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        for reports in executor.map(llm_extractor.extract_reports, candidates):
            out.extend(reports)
    return out


def analyze_asm_reports(parsed_reports: list,
                        deposit_low: int = 2, deposit_high: int = 5) -> dict:
    """Phân tích báo cáo ASM (daily_shop_vt only)."""
    parsed_reports = [r for r in parsed_reports
                      if r.get("report_type") == "daily_shop_vt"
                      and r.get("parse_error") is None]
    all_shops, low_deposit_shops, high_deposit_shops, no_deposit_shops = [], [], [], []
    ideas, tich_cuc_list, han_che_list = [], [], []
    total_deposits = 0
    total_ra_tiem  = 0

    for r in parsed_reports:
        dep = r.get("deposit_count")
        if dep is not None:
            total_deposits += dep
            if dep < deposit_low:
                level = "Thấp"
            elif dep > deposit_high:
                level = "Cao"
            else:
                level = "Bình thường"
            entry = {"shop_ref": r["shop_ref"], "deposit_count": dep,
                     "sender": r["sender"], "level": level}
            all_shops.append(entry)
            if dep == 0:
                no_deposit_shops.append({"sender": r["sender"], "shop_ref": r["shop_ref"]})
            elif level == "Thấp":
                low_deposit_shops.append(entry)
            elif level == "Cao":
                high_deposit_shops.append(entry)

        tiem = r.get("ra_tiem_count")
        if tiem is not None:
            total_ra_tiem += tiem

        if r.get("da_lam"):
            ideas.append({
                "sender": r["sender"], "shop_ref": r["shop_ref"],
                "da_lam": r["da_lam"], "sent_at": r["sent_at"],
            })

        if r.get("tich_cuc"):
            tich_cuc_list.append({
                "sender": r["sender"], "shop_ref": r["shop_ref"], "content": r["tich_cuc"],
            })

        if r.get("van_de"):
            han_che_list.append({
                "sender": r["sender"], "shop_ref": r["shop_ref"], "content": r["van_de"],
            })

    return {
        "total_deposits":    total_deposits,
        "total_ra_tiem":     total_ra_tiem,
        "all_shops":         all_shops,
        "no_deposit_shops":  no_deposit_shops,
        "low_deposit_shops": low_deposit_shops,
        "high_deposit_shops": high_deposit_shops,
        "ideas":             ideas,
        "highlights":        {"tich_cuc": tich_cuc_list, "han_che": han_che_list},
        "missing_reporters": None,  # None = chưa kiểm tra; [] = tất cả đã báo cáo
    }


def analyze_tttc_reports(parsed: list) -> dict:
    """Tổng hợp các TTTC report đã parse. Mọi tỉ số chỉ tính trên non-null."""
    parsed = [r for r in parsed
              if r.get("report_type") == "weekend_tttc"
              and r.get("parse_error") is None]
    def _mean(xs):
        xs = [x for x in xs if x is not None]
        if not xs:
            return None
        return sum(xs) / len(xs)

    avg_tb_bill = _mean([r["tb_bill_vnd"] for r in parsed])
    if avg_tb_bill is not None:
        avg_tb_bill = int(round(avg_tb_bill))

    avg_revenue_pct = _mean([r["revenue_pct"] for r in parsed])
    avg_hot_pct     = _mean([r["hot_pct"]     for r in parsed])
    avg_hot_ratio   = _mean([r["hot_ratio_pct"]   for r in parsed])

    # Nulls-last stable sort
    def _sort_top(r):
        v = r["revenue_pct"]
        return (v is None, -(v or 0))

    def _sort_bottom(r):
        v = r["revenue_pct"]
        return (v is None, (v or 0))

    # Shallow-copy so downstream mutation doesn't alias back into parsed_tttc.
    # Also project the venue = shop_ref alias so downstream callsites that
    # still read r["venue"] keep working.
    def _copy(r):
        d = {**r}
        d["venue"] = r.get("shop_ref")
        return d
    top    = [_copy(r) for r in sorted(parsed, key=_sort_top)[:5]]
    bottom = [_copy(r) for r in sorted(parsed, key=_sort_bottom)[:5]]

    ideas = [
        {"sender": r["sender"], "venue": r.get("shop_ref"),
         "da_lam": r["da_lam"], "sent_at": r.get("sent_at", "")}
        for r in parsed
        if r.get("da_lam")
    ]
    tich_cuc = [
        {"sender": r["sender"], "venue": r.get("shop_ref"), "content": r["tich_cuc"]}
        for r in parsed
        if r.get("tich_cuc")
    ]
    han_che = [
        {"sender": r["sender"], "venue": r.get("shop_ref"), "content": r["van_de"]}
        for r in parsed
        if r.get("van_de")
    ]

    return {
        "total_reports":   len(parsed),
        "avg_tb_bill":     avg_tb_bill,
        "avg_revenue_pct": avg_revenue_pct,
        "avg_hot_pct":     avg_hot_pct,
        "avg_hot_ratio":   avg_hot_ratio,
        "top_centers":     top,
        "bottom_centers":  bottom,
        "ideas":           ideas,
        "highlights": {
            "tich_cuc": tich_cuc,
            "han_che":  han_che,
        },
    }


def report_type_for_date(target_date: _date) -> ReportType:
    """Mon-Fri (weekday 0-4) → Shop VT daily; Sat-Sun (5-6) → TTTC weekend."""
    return "daily_shop_vt" if target_date.weekday() < 5 else "weekend_tttc"


def _is_active_member(m: dict) -> bool:
    """Active = đã từng đọc tin nhắn trong group.

    Loại zombie account (data quality issue: cùng người có 2 user record,
    1 active + 1 zombie với lastReadMessageId=0). Chỉ dùng cho compliance —
    KHÔNG dùng cho display/sender lookup.
    """
    return (m.get("lastReadMessageId") or 0) > 0


def check_asm_compliance(parsed_reports: list, members: list,
                         target_date_str: str,
                         report_type: ReportType,
                         deadline_hhmm: str = "20:00",
                         skip_list: list | None = None) -> list:
    """Trả về displayName của thành viên chưa gửi báo cáo trước deadline.

    `report_type` REQUIRED — caller phải route theo weekday (xem
    `report_type_for_date`). Default sẽ tạo silent bug khi quên route.
    Members có lastReadMessageId=0 bị filter (zombie account).
    """
    parsed_reports = [r for r in parsed_reports
                      if r.get("report_type") == report_type
                      and r.get("parse_error") is None]
    members = [m for m in members if _is_active_member(m)]
    VN_OFFSET = 7 * 3600
    try:
        deadline_h, deadline_m = map(int, deadline_hhmm.split(":"))
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    except ValueError as e:
        print(f"  [!] Tham số không hợp lệ: {e}", file=sys.stderr)
        return []

    reported = set()
    for r in parsed_reports:
        dt = parse_dt(r.get("sent_at", ""))
        if not dt:
            continue
        vn_dt = datetime.fromtimestamp(dt.timestamp() + VN_OFFSET, tz=timezone.utc)
        if vn_dt.date() != target_date:
            continue
        if vn_dt.hour > deadline_h or (vn_dt.hour == deadline_h and vn_dt.minute >= deadline_m):
            continue
        reported.add(r["sender"].strip().lower())

    skip = [s.lower() for s in (skip_list or [])]
    missing = []
    for m in members:
        name = (m.get("displayName") or "").strip()
        if not name:
            continue
        name_lower = name.lower()
        if any(s in name_lower for s in skip):
            continue
        if not any(name_lower in rn or rn in name_lower for rn in reported):
            missing.append(name)
    return missing


def check_late_reporters(parsed_reports: list,
                         target_date_str: str,
                         deadline_hhmm: str = "20:00") -> list:
    """Trả về list {sender, sent_at_vn} của ASM gửi báo cáo SAU deadline."""
    parsed_reports = [r for r in parsed_reports
                      if r.get("report_type") == "daily_shop_vt"
                      and r.get("parse_error") is None]
    VN_OFFSET = 7 * 3600
    try:
        deadline_h, deadline_m = map(int, deadline_hhmm.split(":"))
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    except ValueError:
        return []

    late = []
    seen = set()
    for r in parsed_reports:
        dt = parse_dt(r.get("sent_at", ""))
        if not dt:
            continue
        vn_dt = datetime.fromtimestamp(dt.timestamp() + VN_OFFSET, tz=timezone.utc)
        if vn_dt.date() != target_date:
            continue
        after_deadline = (vn_dt.hour > deadline_h or
                          (vn_dt.hour == deadline_h and vn_dt.minute >= deadline_m))
        if after_deadline and r["sender"] not in seen:
            seen.add(r["sender"])
            late.append({
                "sender":     r["sender"],
                "sent_at_vn": vn_dt.strftime("%H:%M"),
            })
    return late


def analyze_multiday(parsed_reports: list, date_from_str: str, date_to_str: str) -> dict:
    """Phân tích báo cáo ASM theo từng ngày trong khoảng nhiều ngày."""
    parsed_reports = [r for r in parsed_reports
                      if r.get("report_type") == "daily_shop_vt"
                      and r.get("parse_error") is None]
    VN_OFFSET = 7 * 3600

    d_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
    d_to   = datetime.strptime(date_to_str,   "%Y-%m-%d").date()
    total_days = (d_to - d_from).days + 1
    all_dates  = [d_from + timedelta(days=i) for i in range(total_days)]

    # Group reports by (date, sender)
    by_date: dict[_date, list] = {d: [] for d in all_dates}
    for r in parsed_reports:
        dt = parse_dt(r.get("sent_at", ""))
        if not dt:
            continue
        vn_dt = datetime.fromtimestamp(dt.timestamp() + VN_OFFSET, tz=timezone.utc)
        if d_from <= vn_dt.date() <= d_to:
            by_date[vn_dt.date()].append(r)

    # All senders who appeared at least once
    all_senders = sorted({r["sender"] for r in parsed_reports if r.get("sender")})

    # daily_summary
    daily_summary = []
    for d in all_dates:
        reps = by_date[d]
        daily_summary.append({
            "date":           d.strftime("%Y-%m-%d"),
            "total_deposits": sum(r["deposit_count"] for r in reps if r.get("deposit_count") is not None),
            "total_ra_tiem":  sum(r["ra_tiem_count"]  for r in reps if r.get("ra_tiem_count")  is not None),
            "reporter_count": len({r["sender"] for r in reps}),
        })

    # Per-sender reported days set
    sender_dates: dict[str, set] = {s: set() for s in all_senders}
    for d, reps in by_date.items():
        for r in reps:
            if r.get("sender"):
                sender_dates[r["sender"]].add(d)

    # asm_summary with streak calculation
    def _streaks(reported_set: set, all_d: list):
        longest_streak = longest_gap = cur_streak = cur_gap = 0
        for d in all_d:
            if d in reported_set:
                cur_streak += 1
                cur_gap = 0
            else:
                cur_gap += 1
                cur_streak = 0
            longest_streak = max(longest_streak, cur_streak)
            longest_gap    = max(longest_gap,    cur_gap)
        return longest_streak, longest_gap

    asm_summary = []
    for sender in all_senders:
        rep_dates = sender_dates[sender]
        report_days = len(rep_dates)
        report_rate = round(report_days / total_days * 100, 1) if total_days else 0
        sender_reps = [r for r in parsed_reports
                       if r.get("sender") == sender and r.get("deposit_count") is not None]
        total_dep = sum(r["deposit_count"] for r in sender_reps)
        avg_dep   = round(total_dep / report_days, 1) if report_days else 0
        ls, lg    = _streaks(rep_dates, all_dates)
        asm_summary.append({
            "sender":              sender,
            "report_days":         report_days,
            "total_days":          total_days,
            "report_rate":         report_rate,
            "total_deposits":      total_dep,
            "avg_deposits_per_day": avg_dep,
            "longest_streak":      ls,
            "longest_gap":         lg,
        })
    asm_summary.sort(key=lambda x: (-x["report_rate"], -x["total_deposits"]))

    # missing_by_day
    missing_by_day = []
    for d in all_dates:
        reported_today = {r["sender"] for r in by_date[d]}
        missing = sorted(s for s in all_senders if s not in reported_today)
        missing_by_day.append({
            "date":          d.strftime("%Y-%m-%d"),
            "missing_count": len(missing),
            "missing_names": missing,
        })

    # shop_summary
    shop_acc: dict[str, dict] = {}
    for r in parsed_reports:
        shop = r.get("shop_ref")
        if not shop:
            continue
        dt = parse_dt(r.get("sent_at", ""))
        if not dt:
            continue
        vn_dt = datetime.fromtimestamp(dt.timestamp() + VN_OFFSET, tz=timezone.utc)
        if not (d_from <= vn_dt.date() <= d_to):
            continue
        dep = r.get("deposit_count") or 0
        if shop not in shop_acc:
            shop_acc[shop] = {"sender": r.get("sender", ""), "total_deposits": 0, "dates": set()}
        shop_acc[shop]["total_deposits"] += dep
        shop_acc[shop]["dates"].add(vn_dt.date())

    shop_summary = sorted([
        {
            "shop_ref":      shop,
            "sender":        v["sender"],
            "total_deposits": v["total_deposits"],
            "report_days":   len(v["dates"]),
            "avg_deposits":  round(v["total_deposits"] / len(v["dates"]), 1) if v["dates"] else 0,
        }
        for shop, v in shop_acc.items()
    ], key=lambda x: -x["total_deposits"])

    return {
        "total_days":     total_days,
        "daily_summary":  daily_summary,
        "asm_summary":    asm_summary,
        "missing_by_day": missing_by_day,
        "shop_summary":   shop_summary,
    }


def analyze_weekly(messages: list,
                   group_members: list,
                   target_date_vn: str,
                   deadline: str = "20:00") -> dict:
    """Báo cáo tuần một ngày: phân loại tin nhắn, tìm muộn/thiếu, dump text."""
    VN_OFFSET = 7 * 3600
    target = datetime.strptime(target_date_vn, "%Y-%m-%d").date()
    dl_h, dl_m = [int(x) for x in deadline.split(":", 1)]
    deadline_time = time(hour=dl_h, minute=dl_m)

    member_names = sorted({
        (m.get("displayName") or "").strip()
        for m in group_members
        if (m.get("displayName") or "").strip()
    })
    member_set = set(member_names)

    # Group qualifying messages by sender
    by_sender: dict[str, list] = {}
    for msg in messages:
        if msg.get("type") != "TEXT":
            continue
        content = msg.get("content") or ""
        if not content.strip():
            continue
        user = msg.get("user") or {}
        sender = (user.get("displayName") or "").strip()
        if sender not in member_set:
            continue
        dt = parse_dt(msg.get("createdAt", ""))
        if not dt:
            continue
        vn_dt = datetime.fromtimestamp(dt.timestamp() + VN_OFFSET, tz=timezone.utc)
        if vn_dt.date() != target:
            continue
        if _score_weekly_message(content) < WEEKLY_SCORE_THRESHOLD:
            continue
        by_sender.setdefault(sender, []).append((dt, vn_dt, content))

    reports = []
    for sender, items in by_sender.items():
        items.sort(key=lambda t: t[0])
        _, vn_dt, content = items[0]
        is_late = vn_dt.time() >= deadline_time
        reports.append({
            "sender": sender,
            "sent_at_vn": vn_dt.strftime("%H:%M"),
            "is_late": is_late,
            "text": content,
            "extra_count": len(items) - 1,
        })
    reports.sort(key=lambda r: r["sent_at_vn"])

    late_list = sorted(r["sender"] for r in reports if r["is_late"])
    reporters = {r["sender"] for r in reports}
    missing_list = sorted(n for n in member_names if n not in reporters)

    # --- Structured dispatch: LLM extraction, split by report_type ---
    import llm_extractor
    parsed_shop_vt: list = []
    parsed_tttc:    list = []
    unparseable:    list = []
    for sender, items in by_sender.items():
        for dt, _vn_dt, content in items:
            fake_msg = {
                "content":   content,
                "user":      {"displayName": sender, "id": ""},
                "createdAt": dt.isoformat(),
                "id":        f"{sender}-{dt.timestamp()}",
                "type":      "TEXT",
            }
            for r in llm_extractor.extract_reports(fake_msg):
                if r.get("parse_error") is not None:
                    unparseable.append(r)
                elif r["report_type"] == "daily_shop_vt":
                    parsed_shop_vt.append(r)
                elif r["report_type"] == "weekend_tttc":
                    parsed_tttc.append(r)

    asm_data  = analyze_asm_reports(parsed_shop_vt) if parsed_shop_vt else None
    tttc_data = analyze_tttc_reports(parsed_tttc)   if parsed_tttc    else None

    return {
        "target_date":    target_date_vn,
        "deadline":       deadline,
        "reports":        reports,
        "late_list":      late_list,
        "missing_list":   missing_list,
        "asm_data":       asm_data,
        "tttc_data":      tttc_data,
        "parsed_shop_vt": parsed_shop_vt,
        "parsed_tttc":    parsed_tttc,
        "unparseable":    unparseable,
    }


# ---------------------------------------------------------------------------
# Report Output
# ---------------------------------------------------------------------------

def print_asm_report(asm_data: dict,
                     date_from: datetime | None = None,
                     date_to: datetime | None = None) -> None:
    sep = "=" * 65
    print(sep)
    print("  FPT CHAT - BÁO CÁO ASM")
    print(sep)

    print(f"\n{'TỔNG QUAN':=<40}")
    print(f"  Khoảng thời gian : {fmt_date_range(date_from, date_to)}")
    print(f"  Báo cáo ASM      : {len(asm_data.get('all_shops', []))}")

    low_shops  = asm_data.get("low_deposit_shops", [])
    high_shops = asm_data.get("high_deposit_shops", [])
    ideas      = asm_data.get("ideas", [])
    tich_cuc   = asm_data.get("highlights", {}).get("tich_cuc", [])
    han_che    = asm_data.get("highlights", {}).get("han_che", [])
    missing    = asm_data.get("missing_reporters")

    print(f"\n{'SHOP ĐẶT CỌC THẤP (< ngưỡng)':=<40}")
    if low_shops:
        for s in sorted(low_shops, key=lambda x: x["deposit_count"]):
            print(f"  [{s['deposit_count']:>3} đặt cọc] {s['shop_ref']}  — {s['sender']}")
    else:
        print("  (không có)")

    print(f"\n{'SHOP ĐẶT CỌC CAO (> ngưỡng)':=<40}")
    if high_shops:
        for s in sorted(high_shops, key=lambda x: x["deposit_count"], reverse=True):
            print(f"  [{s['deposit_count']:>3} đặt cọc] {s['shop_ref']}  — {s['sender']}")
    else:
        print("  (không có)")

    print(f"\n{'Ý TƯỞNG TRIỂN KHAI TỪ ASM':=<40}")
    if ideas:
        for i, idea in enumerate(ideas, 1):
            print(f"  {i}. [{idea['sender']}] Shop: {idea['shop_ref']}")
            for line in idea["da_lam"].splitlines():
                if line.strip():
                    print(f"     {line.strip()}")
    else:
        print("  (không có)")

    print(f"\n{'ĐIỂM TÍCH CỰC':=<40}")
    if tich_cuc:
        for i, h in enumerate(tich_cuc, 1):
            print(f"  {i}. [{h['sender']}] Shop: {h['shop_ref']}")
            for line in h["content"].splitlines():
                if line.strip():
                    print(f"     {line.strip()}")
    else:
        print("  (không có)")

    print(f"\n{'ĐIỂM HẠN CHẾ':=<40}")
    if han_che:
        for i, h in enumerate(han_che, 1):
            print(f"  {i}. [{h['sender']}] Shop: {h['shop_ref']}")
            for line in h["content"].splitlines():
                if line.strip():
                    print(f"     {line.strip()}")
    else:
        print("  (không có)")

    if missing is not None:
        print(f"\n{'ASM CHƯA BÁO CÁO':=<40}")
        if missing:
            for name in missing:
                print(f"  - {name}")
        else:
            print("  Tất cả ASM đã báo cáo")

    print(f"\n{sep}")


def print_weekly_report(data: dict) -> None:
    """In báo cáo tuần một ngày ra stdout."""
    target = data["target_date"]
    deadline = data["deadline"]
    reports = data["reports"]
    late_list = data["late_list"]
    missing_list = data["missing_list"]

    sep = "=" * 65
    print(sep)
    print(f"  BÁO CÁO TUẦN — {target}")
    print(sep)
    print(f"  Deadline: {deadline}")
    print(f"  Đã báo cáo: {len(reports)}")
    print(f"  Muộn: {len(late_list)}")
    print(f"  Chưa báo cáo: {len(missing_list)}")
    print()

    if missing_list:
        print(f"--- Chưa báo cáo ({len(missing_list)}) ---")
        for name in missing_list:
            print(f"  - {name}")
        print()

    if late_list:
        print(f"--- Muộn ({len(late_list)}) ---")
        late_map = {r["sender"]: r["sent_at_vn"] for r in reports if r["is_late"]}
        for name in late_list:
            print(f"  - {name} ({late_map.get(name, '?')})")
        print()

    asm_data  = data.get("asm_data")
    tttc_data = data.get("tttc_data")

    if asm_data:
        print(f"--- Shop VT: tổng cọc {asm_data['total_deposits']}, "
              f"tổng ra tiêm {asm_data['total_ra_tiem']} ---")
        low  = asm_data.get("low_deposit_shops", [])
        high = asm_data.get("high_deposit_shops", [])
        if low:
            print(f"  Shop cọc thấp ({len(low)}):")
            for s in low:
                print(f"    - {s['shop_ref']} — {s['deposit_count']} cọc ({s['sender']})")
        if high:
            print(f"  Shop cọc tốt ({len(high)}):")
            for s in high:
                print(f"    - {s['shop_ref']} — {s['deposit_count']} cọc ({s['sender']})")
        print()

    if tttc_data:
        def _fmt_pct(v): return f"{v:.1f}%" if v is not None else "—"
        def _fmt_vnd(v): return f"{v:,} đ" if v is not None else "—"
        print(f"--- TTTC: {tttc_data['total_reports']} trung tâm ---")
        print(f"  TB bill TB : {_fmt_vnd(tttc_data['avg_tb_bill'])}")
        print(f"  %HT TB     : {_fmt_pct(tttc_data['avg_revenue_pct'])}")
        print(f"  %HOT TB    : {_fmt_pct(tttc_data['avg_hot_pct'])}")
        print(f"  %tr HOT TB : {_fmt_pct(tttc_data['avg_hot_ratio'])}")
        if tttc_data["top_centers"]:
            print(f"  Top trung tâm theo %HT:")
            for c in tttc_data["top_centers"]:
                print(f"    - {c['venue']} — %HT {_fmt_pct(c['revenue_pct'])} "
                      f"({c['sender']})")
        if tttc_data["bottom_centers"]:
            print(f"  Trung tâm cần chú ý:")
            for c in tttc_data["bottom_centers"]:
                print(f"    - {c['venue']} — %HT {_fmt_pct(c['revenue_pct'])} "
                      f"({c['sender']})")
        print()

    if reports:
        print(f"--- Nội dung báo cáo ({len(reports)}) ---")
        for r in reports:
            suffix = " — MUỘN" if r["is_late"] else ""
            extra  = f" (+{r['extra_count']} tin nhắn khác)" if r["extra_count"] else ""
            print(f"[{r['sender']} — {r['sent_at_vn']}{suffix}{extra}]")
            print(r["text"])
            print()


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def write_asm_excel(asm_data: dict, path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Thiếu thư viện 'openpyxl'. Chạy: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def set_widths(ws, widths: dict):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def wrap_col(ws, col: int):
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    wb = Workbook()

    # ── Sheet 1: Shop Đặt Cọc ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Shop Đặt Cọc"
    ws1.append(["STT", "Shop", "Số đặt cọc", "Ra tiêm", "Mức", "ASM"])
    style_header(ws1)
    all_shops = sorted(asm_data.get("all_shops", []),
                       key=lambda x: x["deposit_count"], reverse=True)
    # Build ra_tiem lookup by shop_ref from parsed_reports if available
    _ra_tiem_map = {r["shop_ref"]: r.get("ra_tiem_count")
                    for r in asm_data.get("parsed_reports", []) if r.get("shop_ref")}
    for i, s in enumerate(all_shops, 1):
        ra = _ra_tiem_map.get(s["shop_ref"], "")
        ws1.append([i, s["shop_ref"], s["deposit_count"],
                    "" if ra is None else ra, s["level"], s["sender"]])
    set_widths(ws1, {1: 6, 2: 50, 3: 14, 4: 12, 5: 14, 6: 35})

    # ── Sheet 2: Ý tưởng ASM ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Ý tưởng ASM")
    ws2.append(["STT", "ASM", "Shop", "Nội dung", "Ngày giờ (UTC+7)"])
    style_header(ws2)
    for i, idea in enumerate(asm_data.get("ideas", []), 1):
        ws2.append([i, idea["sender"], idea["shop_ref"],
                    idea["da_lam"], to_vn_str(idea["sent_at"])])
    set_widths(ws2, {1: 6, 2: 30, 3: 40, 4: 80, 5: 22})
    wrap_col(ws2, 4)

    # ── Sheet 3: Điểm nổi bật ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Điểm nổi bật")
    ws3.append(["STT", "ASM", "Shop", "Loại", "Nội dung"])
    style_header(ws3)
    highlights = (
        [(h["sender"], h["shop_ref"], "Tích cực", h["content"])
         for h in asm_data.get("highlights", {}).get("tich_cuc", [])]
        + [(h["sender"], h["shop_ref"], "Hạn chế", h["content"])
           for h in asm_data.get("highlights", {}).get("han_che", [])]
    )
    for i, (sender, shop, loai, content) in enumerate(highlights, 1):
        ws3.append([i, sender, shop, loai, content])
    set_widths(ws3, {1: 6, 2: 30, 3: 40, 4: 12, 5: 80})
    wrap_col(ws3, 5)

    # ── Sheet 4: Báo cáo muộn ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Báo cáo muộn")
    ws4.append(["STT", "ASM", "Giờ gửi (UTC+7)"])
    style_header(ws4)
    late = asm_data.get("late_reporters", [])
    if late:
        for i, lr in enumerate(late, 1):
            ws4.append([i, lr["sender"], lr["sent_at_vn"]])
    else:
        ws4.append(["", "Không có ASM báo cáo muộn"])
    set_widths(ws4, {1: 6, 2: 45, 3: 18})

    # ── Sheet 5: Chưa báo cáo sau deadline ───────────────────────────────
    ws5 = wb.create_sheet("Chưa báo cáo sau deadline")
    ws5.append(["STT", "Tên ASM"])
    style_header(ws5)
    missing = asm_data.get("missing_reporters")
    if missing is None:
        ws5.append(["", "(Không thể kiểm tra — thiếu token/group)"])
    elif not missing:
        ws5.append(["", "Tất cả ASM đã báo cáo đúng hạn"])
    else:
        for i, name in enumerate(missing, 1):
            ws5.append([i, name])
    set_widths(ws5, {1: 6, 2: 45})

    # ── Sheet 6: Chưa báo cáo tại thời điểm chạy ─────────────────────────
    ws6 = wb.create_sheet("Chưa báo cáo hiện tại")
    ws6.append(["STT", "Tên thành viên"])
    style_header(ws6)
    unreported = asm_data.get("unreported_now")
    if unreported is None:
        ws6.append(["", "(Không thể kiểm tra — thiếu token/group)"])
    elif not unreported:
        ws6.append(["", "Tất cả đã báo cáo"])
    else:
        for i, name in enumerate(unreported, 1):
            ws6.append([i, name])
    set_widths(ws6, {1: 6, 2: 45})

    wb.save(path)


def write_weekly_excel(data: dict, group_members: list, path) -> None:
    """Xuất báo cáo tuần ra .xlsx với 4 sheet: Tổng hợp tuần, Nội dung, Shop VT, TTTC."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("Thiếu 'openpyxl'. Chạy: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()

    # Sheet 1: Tổng hợp tuần
    ws1 = wb.active
    ws1.title = "Tổng hợp tuần"
    ws1.append(["Người báo cáo", "Trạng thái", "Giờ gửi"])

    reports_by_sender = {r["sender"]: r for r in data["reports"]}
    member_names = sorted({
        (m.get("displayName") or "").strip()
        for m in group_members
        if (m.get("displayName") or "").strip()
    })

    def _row_for(name: str):
        r = reports_by_sender.get(name)
        if r is None:
            return (name, "Chưa báo cáo", "")
        return (name, "Muộn" if r["is_late"] else "Đúng giờ", r["sent_at_vn"])

    rows = [_row_for(n) for n in member_names]
    status_order = {"Đúng giờ": 0, "Muộn": 1, "Chưa báo cáo": 2}
    rows.sort(key=lambda r: (status_order[r[1]], r[2] or "ZZ", r[0]))
    for row in rows:
        ws1.append(list(row))

    for cell in ws1[1]:
        cell.font = Font(bold=True)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 10

    # Sheet 2: Nội dung
    ws2 = wb.create_sheet("Nội dung")
    ws2.append(["Người báo cáo", "Giờ gửi", "Trạng thái", "Nội dung"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for r in sorted(data["reports"], key=lambda r: r["sent_at_vn"]):
        status = "Muộn" if r["is_late"] else "Đúng giờ"
        extra = f"\n\n(+{r['extra_count']} tin nhắn khác)" if r["extra_count"] else ""
        body = r["text"] + extra
        ws2.append([r["sender"], r["sent_at_vn"], status, body])

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 100

    for row_idx in range(2, ws2.max_row + 1):
        ws2.cell(row=row_idx, column=4).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    # Sheet 3: Shop VT
    ws3 = wb.create_sheet("Shop VT")
    ws3.append(["Shop", "Số cọc", "Ra tiêm", "Mức", "ASM"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    asm_data = data.get("asm_data") or {}
    _rt_map = {r["shop_ref"]: r.get("ra_tiem_count")
               for r in data.get("parsed_shop_vt", []) if r.get("shop_ref")}
    for s in sorted(asm_data.get("all_shops", []),
                    key=lambda x: x["deposit_count"], reverse=True):
        ra = _rt_map.get(s["shop_ref"])
        ws3.append([s["shop_ref"], s["deposit_count"],
                    "" if ra is None else ra, s["level"], s["sender"]])
    ws3.column_dimensions["A"].width = 50
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 28

    # Sheet 4: TTTC
    ws4 = wb.create_sheet("TTTC")
    ws4.append(["Trung tâm", "%HT ngày", "%HOT", "TB bill",
                "Tỉ trọng HOT", "Lượt KH mua", "ASM"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for r in data.get("parsed_tttc", []):
        ws4.append([
            r.get("venue") or "",
            r.get("revenue_pct")    if r.get("revenue_pct")    is not None else "",
            r.get("hot_pct")        if r.get("hot_pct")        is not None else "",
            r.get("tb_bill")        if r.get("tb_bill")        is not None else "",
            r.get("hot_ratio")      if r.get("hot_ratio")      is not None else "",
            r.get("customer_count") if r.get("customer_count") is not None else "",
            r.get("sender") or "",
        ])
    ws4.column_dimensions["A"].width = 40
    for col in ("B", "C", "D", "E", "F"):
        ws4.column_dimensions[col].width = 14
    ws4.column_dimensions["G"].width = 28

    wb.save(path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    pre = argparse.ArgumentParser(add_help=False)
    pre.add_argument("--config", default="config.json")
    pre_args, _ = pre.parse_known_args()
    cfg = load_config(pre_args.config)

    parser = argparse.ArgumentParser(
        description="FPT Chat ASM Report Tool — phân tích báo cáo hàng ngày của ASM",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python fpt_chat_stats.py --today
  python fpt_chat_stats.py --today --excel bao_cao.xlsx
  python fpt_chat_stats.py --from 2026-04-01 --to 2026-04-16
  python fpt_chat_stats.py --save raw.json
  python fpt_chat_stats.py --load raw.json --today
        """,
    )
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--token", default=None,
                        help="Token xác thực (fchat_ddtk). Ưu tiên hơn config file.")
    parser.add_argument("--group", default=None,
                        help="Group ID hoặc URL chat group.")
    parser.add_argument("--api-url", default=None,
                        help="Base URL của API (mặc định: https://api-chat.fpt.com)")
    parser.add_argument("--limit", type=int, default=50,
                        help="Số tin nhắn mỗi trang API (mặc định: 50)")
    parser.add_argument("--from", dest="date_from", default=None, metavar="YYYY-MM-DD",
                        help="Chỉ phân tích từ ngày này (inclusive)")
    parser.add_argument("--to", dest="date_to", default=None, metavar="YYYY-MM-DD",
                        help="Chỉ phân tích đến ngày này (inclusive)")
    parser.add_argument("--today", action="store_true",
                        help="Phân tích hôm nay (giờ VN). Không dùng kèm --from/--to/--date.")
    parser.add_argument("--save", metavar="FILE",
                        help="Lưu raw messages ra file JSON")
    parser.add_argument("--load", metavar="FILE",
                        help="Dùng file JSON đã lưu (offline)")
    parser.add_argument("--excel", metavar="FILE",
                        help="Xuất báo cáo Excel (.xlsx) với 4 sheet ASM")
    parser.add_argument("--deposit-low", dest="deposit_low",
                        type=int, default=cfg.get("asm_deposit_low", 2), metavar="N",
                        help="Ngưỡng đặt cọc thấp — shop có deposit < N (mặc định: 2)")
    parser.add_argument("--deposit-high", dest="deposit_high",
                        type=int, default=cfg.get("asm_deposit_high", 5), metavar="N",
                        help="Ngưỡng đặt cọc cao — shop có deposit > N (mặc định: 5)")
    parser.add_argument("--asm-deadline", default="20:00", metavar="HH:MM",
                        help="Deadline báo cáo hàng ngày (mặc định: 20:00, giờ VN)")
    parser.add_argument("--weekly", default=None, metavar="YYYY-MM-DD",
                        help="Báo cáo tuần một ngày: liệt kê ai đã/muộn/chưa báo cáo + dump text.")
    parser.add_argument("--date", default=None, metavar="YYYY-MM-DD",
                        help="Ngày kiểm tra compliance (mặc định: hôm nay giờ VN)")
    parser.add_argument("--skip-reporters", default="", metavar="NAMES",
                        help="Tên cách nhau bằng dấu phẩy — loại khỏi compliance check")
    parser.add_argument("--llm-base-url", dest="llm_base_url", default=None,
                        help="OpenAI-compatible base URL (default: https://api.openai.com/v1)")
    parser.add_argument("--llm-model", dest="llm_model", default=None,
                        help="LLM model name (default: gpt-5.4-mini)")
    parser.add_argument("--llm-structured-outputs",
                        dest="llm_structured_outputs",
                        action="store_true", default=None,
                        help="Use OpenAI structured outputs (json_schema, strict). "
                             "Only on providers that support it.")
    parser.add_argument("--no-llm-structured-outputs",
                        dest="llm_structured_outputs",
                        action="store_false",
                        help="Force JSON mode even if config enables structured outputs.")
    parser.add_argument("--llm-max-workers", dest="llm_max_workers",
                        type=int, default=None, metavar="N",
                        help="Concurrent LLM calls during extraction (default: 4, "
                             "sized for OpenAI tier 1). Set 1 to disable parallelism.")

    args = parser.parse_args()

    import llm_extractor
    _llm_cfg = cfg.get("llm") or {}
    _so = args.llm_structured_outputs
    if _so is None and "structured_outputs" in _llm_cfg:
        _so = bool(_llm_cfg["structured_outputs"])
    _mw = args.llm_max_workers
    if _mw is None and "max_workers" in _llm_cfg:
        _mw = int(_llm_cfg["max_workers"])
    llm_extractor.configure(
        api_key = _llm_cfg.get("api_key"),
        base_url= args.llm_base_url or _llm_cfg.get("base_url"),
        model   = args.llm_model    or _llm_cfg.get("model"),
        structured_outputs = _so,
        max_workers = _mw,
    )

    if args.weekly and (args.today or args.date_from or args.date_to or args.date):
        print("Error: --weekly không dùng chung với --today / --from / --to / --date.", file=sys.stderr)
        sys.exit(2)

    token   = args.token   or cfg.get("token")
    group   = args.group   or cfg.get("group")
    api_url = args.api_url or cfg.get("api_url", "https://api-chat.fpt.com")

    # ── Validate (chỉ khi cần gọi API; --weekly tự kiểm tra bên trong)
    if not args.load and not args.weekly:
        if not token:
            print("Error: --token là bắt buộc (hoặc set 'token' trong config.json)", file=sys.stderr)
            sys.exit(1)
        if not group:
            print("Error: --group là bắt buộc (hoặc set 'group' trong config.json)", file=sys.stderr)
            sys.exit(1)

    # ── --today shortcut
    if args.today:
        if args.date_from or args.date_to or args.date:
            print("Error: --today không thể dùng kèm --from, --to, hoặc --date", file=sys.stderr)
            sys.exit(1)
        import time as _time
        _vn_today = datetime.fromtimestamp(
            _time.time() + 7 * 3600, tz=timezone.utc
        ).strftime("%Y-%m-%d")
        args.date_from = _vn_today
        args.date_to   = _vn_today
        args.date      = _vn_today

    # ── --weekly shortcut (báo cáo tuần một ngày)
    if args.weekly:
        try:
            target_vn = datetime.strptime(args.weekly, "%Y-%m-%d").date()
        except ValueError:
            print(f"Error: --weekly giá trị không hợp lệ: {args.weekly}", file=sys.stderr)
            sys.exit(2)

        # Half-open VN-day window: [target 00:00+07, target+1 00:00+07)
        vn_start_utc = datetime.combine(target_vn, time(0, 0), tzinfo=timezone.utc) - timedelta(hours=7)
        vn_end_utc   = vn_start_utc + timedelta(days=1)

        # Fetch or load raw messages
        if args.load:
            print(f"[*] Loading từ file: {args.load}", file=sys.stderr)
            with open(args.load, encoding="utf-8") as f:
                messages = json.load(f)
            print(f"[✓] Loaded {len(messages)} messages", file=sys.stderr)
        else:
            if not token:
                print("Error: --weekly cần --token (hoặc config.json).", file=sys.stderr)
                sys.exit(1)
            if not group:
                print("Error: --weekly cần --group (hoặc config.json).", file=sys.stderr)
                sys.exit(1)
            group_id = extract_group_id(group)
            messages = fetch_all_messages(
                token=token, group_id=group_id, base_url=api_url,
                limit=args.limit, date_from=vn_start_utc,
            )

        if args.save:
            with open(args.save, "w", encoding="utf-8") as f:
                json.dump(messages, f, ensure_ascii=False, indent=2)
            print(f"[✓] Đã lưu raw messages → {args.save}", file=sys.stderr)

        # Clip to the half-open window. filter_by_date uses inclusive date_to,
        # so subtract 1 microsecond to emulate half-open [start, end).
        messages = filter_by_date(
            messages, vn_start_utc, vn_end_utc - timedelta(microseconds=1),
        )

        # Members are REQUIRED for the missing list.
        if not token or not group:
            print(
                "Error: --weekly cần token+group để lấy thành viên group cho compliance check.",
                file=sys.stderr,
            )
            sys.exit(3)
        _session = build_session(token)
        members = fetch_group_members(_session, api_url, extract_group_id(group))
        if not members:
            print(
                "Error: fetch_group_members trả rỗng hoặc lỗi — hủy báo cáo tuần.",
                file=sys.stderr,
            )
            sys.exit(3)

        deadline = cfg.get("deadline", "20:00")
        data = analyze_weekly(messages, members, args.weekly, deadline=deadline)
        print_weekly_report(data)
        if args.excel:
            write_weekly_excel(data, members, args.excel)
            print(f"[✓] Đã xuất Excel → {args.excel}", file=sys.stderr)
        return

    # ── Skip reporters: merge config + CLI
    skip_list = list(cfg.get("asm_skip_reporters", []))
    if args.skip_reporters:
        skip_list += [n.strip() for n in args.skip_reporters.split(",") if n.strip()]

    date_from = parse_date_arg(args.date_from, end_of_day=False) if args.date_from else None
    date_to   = parse_date_arg(args.date_to,   end_of_day=True)  if args.date_to   else None

    # ── Fetch messages
    if args.load:
        print(f"[*] Loading từ file: {args.load}", file=sys.stderr)
        with open(args.load, encoding="utf-8") as f:
            messages = json.load(f)
        print(f"[✓] Loaded {len(messages)} messages", file=sys.stderr)
    else:
        group_id = extract_group_id(group)
        messages = fetch_all_messages(token=token, group_id=group_id,
                                      base_url=api_url, limit=args.limit,
                                      date_from=date_from)

    # ── Save raw
    if args.save:
        with open(args.save, "w", encoding="utf-8") as f:
            json.dump(messages, f, ensure_ascii=False, indent=2)
        print(f"[✓] Đã lưu raw messages → {args.save}", file=sys.stderr)

    # ── Filter by date range
    messages = filter_by_date(messages, date_from, date_to)

    # ── ASM analysis (always) — LLM extraction handles classification + extract
    parsed_reports = extract_all_reports(messages)
    print(f"[*] Trích xuất {len(parsed_reports)} báo cáo (qua LLM)", file=sys.stderr)
    if not parsed_reports:
        print("  [!] Không tìm thấy báo cáo ASM nào.", file=sys.stderr)

    asm_data = analyze_asm_reports(parsed_reports,
                                   deposit_low=args.deposit_low,
                                   deposit_high=args.deposit_high)

    # ── Compliance check
    if token and group:
        _session = build_session(token)
        members = fetch_group_members(_session, api_url, extract_group_id(group))
    else:
        print("[!] Thiếu --token/--group — bỏ qua kiểm tra compliance", file=sys.stderr)
        members = []

    if members:
        import time as _time
        target_date = args.date or datetime.fromtimestamp(
            _time.time() + 7 * 3600, tz=timezone.utc
        ).strftime("%Y-%m-%d")
        rtype = report_type_for_date(
            datetime.strptime(target_date, "%Y-%m-%d").date()
        )
        asm_data["missing_reporters"] = check_asm_compliance(
            parsed_reports, members, target_date,
            report_type=rtype,
            deadline_hhmm=args.asm_deadline,
            skip_list=skip_list,
        )

    # ── Output
    print_asm_report(asm_data, date_from, date_to)

    if args.excel:
        write_asm_excel(asm_data, args.excel)
        print(f"[✓] Đã xuất Excel → {args.excel}", file=sys.stderr)

    print(llm_extractor.format_stats(), file=sys.stderr)


if __name__ == "__main__":
    main()
