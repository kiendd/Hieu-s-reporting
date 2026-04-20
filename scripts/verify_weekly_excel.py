#!/usr/bin/env python3
"""Verify the weekly Excel writer output shape."""
from __future__ import annotations
import pathlib
import sys
import tempfile

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))
from fpt_chat_stats import analyze_weekly, write_weekly_excel

FAIL = 0
def check(label, cond):
    global FAIL
    status = "PASS" if cond else "FAIL"
    if not cond:
        FAIL += 1
    print(f"  [{status}] {label}")

REPORT = (
    "Dạ em gửi báo cáo đánh giá TTTC 58149\n"
    "- Kết quả: Doanh thu 133% HT, TB bill 2.2M\n"
    "- Tích cực: TVV upsale tốt\n"
    "- Đã làm: hướng dẫn kịch bản\n"
    "Em cảm ơn ạ."
)

def msg(sender, utc_ts, text):
    return {
        "id": f"id-{sender}-{utc_ts}",
        "type": "TEXT",
        "content": text,
        "createdAt": utc_ts,
        "user": {"displayName": sender},
    }

members = [{"displayName": n} for n in ["A", "B", "C", "D", "E"]]
messages = [
    msg("A", "2026-04-20T02:00:00Z", REPORT),
    msg("B", "2026-04-20T14:30:00Z", REPORT),
]
data = analyze_weekly(messages, members, "2026-04-20", deadline="20:00")

with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
    path = tmp.name

write_weekly_excel(data, members, path)

import openpyxl
wb = openpyxl.load_workbook(path)
check("sheet 'Tổng hợp tuần' exists", "Tổng hợp tuần" in wb.sheetnames)
check("sheet 'Nội dung' exists", "Nội dung" in wb.sheetnames)
check("exactly 4 sheets", len(wb.sheetnames) == 4)

ws1 = wb["Tổng hợp tuần"]
header1 = [c.value for c in ws1[1]]
check("Tổng hợp tuần headers", header1 == ["Người báo cáo", "Trạng thái", "Giờ gửi"])
rows1 = [[c.value for c in row] for row in ws1.iter_rows(min_row=2, values_only=False)]
check("5 data rows (one per member)", len(rows1) == 5)

statuses = [r[1] for r in rows1]
check(
    "status order: Đúng giờ → Muộn → Chưa báo cáo",
    statuses == ["Đúng giờ", "Muộn", "Chưa báo cáo", "Chưa báo cáo", "Chưa báo cáo"],
)
missing_hours = [r[2] for r in rows1 if r[1] == "Chưa báo cáo"]
check("missing Giờ gửi is empty/None", all(h in (None, "") for h in missing_hours))

ws2 = wb["Nội dung"]
header2 = [c.value for c in ws2[1]]
check("Nội dung headers", header2 == ["Người báo cáo", "Giờ gửi", "Trạng thái", "Nội dung"])
rows2 = [[c.value for c in row] for row in ws2.iter_rows(min_row=2, values_only=False)]
check("Nội dung has 2 data rows", len(rows2) == 2)
check("Nội dung ordered by Giờ gửi asc", rows2[0][1] < rows2[1][1])

content_cell = ws2.cell(row=2, column=4)
check("Nội dung column has wrap_text enabled",
      content_cell.alignment is not None and content_cell.alignment.wrap_text is True)

# --- New: structured sheets ---
SHOP_VT_BODY = pathlib.Path(
    pathlib.Path(__file__).resolve().parent.parent / "templates/weekend/7"
).read_text(encoding="utf-8")
TTTC_BODY = pathlib.Path(
    pathlib.Path(__file__).resolve().parent.parent / "templates/weekend/1"
).read_text(encoding="utf-8")
MIXED_MSGS = [
    msg("ASM-A", "2026-04-20T02:00:00Z", SHOP_VT_BODY),
    msg("ASM-B", "2026-04-20T03:00:00Z", TTTC_BODY),
]
mixed_data = analyze_weekly(MIXED_MSGS,
                            [{"displayName": "ASM-A"}, {"displayName": "ASM-B"}],
                            "2026-04-20")
import io
buf = io.BytesIO()
write_weekly_excel(mixed_data,
                   [{"displayName": "ASM-A"}, {"displayName": "ASM-B"}],
                   buf)
buf.seek(0)
from openpyxl import load_workbook
wb = load_workbook(buf)
check("Shop VT sheet exists",  "Shop VT" in wb.sheetnames)
check("TTTC sheet exists",     "TTTC"    in wb.sheetnames)
ws_sv = wb["Shop VT"]
ws_tt = wb["TTTC"]
shop_headers = [c.value for c in ws_sv[1]]
check("Shop VT header Shop",    "Shop"    in shop_headers)
check("Shop VT header Số cọc",  "Số cọc"  in shop_headers)
tt_headers = [c.value for c in ws_tt[1]]
check("TTTC header Trung tâm",   "Trung tâm"   in tt_headers)
check("TTTC header %HT ngày",    "%HT ngày"    in tt_headers)
check("TTTC header TB bill",     "TB bill"     in tt_headers)
check("Shop VT has at least one data row",  ws_sv.max_row >= 2)
check("TTTC has at least one data row",     ws_tt.max_row >= 2)

# Edge case: zero reports — sheets still exist but empty (headers only)
empty_data = analyze_weekly([],
                            [{"displayName": "ASM-A"}],
                            "2026-04-20")
buf2 = io.BytesIO()
write_weekly_excel(empty_data, [{"displayName": "ASM-A"}], buf2)
buf2.seek(0)
wb2 = load_workbook(buf2)
check("zero-day: Shop VT sheet present", "Shop VT" in wb2.sheetnames)
check("zero-day: TTTC sheet present",    "TTTC"    in wb2.sheetnames)

print(f"\n{FAIL} failure(s)" if FAIL else "\nAll checks passed.")
pathlib.Path(path).unlink(missing_ok=True)
sys.exit(1 if FAIL else 0)
